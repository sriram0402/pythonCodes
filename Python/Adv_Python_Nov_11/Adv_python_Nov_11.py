Python 3.10.4 (v3.10.4:9d38120e33, Mar 23 2022, 17:29:05) [Clang 13.0.0 (clang-1300.0.29.30)] on darwin
Type "help", "copyright", "credits" or "license()" for more information.
import openpyxl
wb = openpyxl.load_workbook(filename='sample3.xlsx')
Traceback (most recent call last):
  File "<pyshell#1>", line 1, in <module>
    wb = openpyxl.load_workbook(filename='sample3.xlsx')
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/reader/excel.py", line 315, in load_workbook
    reader = ExcelReader(filename, read_only, keep_vba,
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/reader/excel.py", line 124, in __init__
    self.archive = _validate_archive(fn)
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/reader/excel.py", line 96, in _validate_archive
    archive = ZipFile(filename, 'r')
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/zipfile.py", line 1249, in __init__
    self.fp = io.open(file, filemode)
FileNotFoundError: [Errno 2] No such file or directory: 'sample3.xlsx'
wb = openpyxl.load_workbook(filename='Test2.xlsx')
                         
Traceback (most recent call last):
  File "<pyshell#2>", line 1, in <module>
    wb = openpyxl.load_workbook(filename='Test2.xlsx')
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/reader/excel.py", line 315, in load_workbook
    reader = ExcelReader(filename, read_only, keep_vba,
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/reader/excel.py", line 124, in __init__
    self.archive = _validate_archive(fn)
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/reader/excel.py", line 96, in _validate_archive
    archive = ZipFile(filename, 'r')
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/zipfile.py", line 1249, in __init__
    self.fp = io.open(file, filemode)
FileNotFoundError: [Errno 2] No such file or directory: 'Test2.xlsx'
wb=openpyxl.Workbook()
                         
ws = wb.create_sheet('Test 2')
                         
wb.save(filename='Sample.xlsx')
                         
wb = openpyxl.load_workbook(filename='Sample.xlsx')
                         
print(ws.cell(row=4, column=3).value)
                         
None
for i in range(1,wd.max_row+1):
                         v=ws.cell(row=i,column=1).value='12'

                         
Traceback (most recent call last):
  File "<pyshell#11>", line 1, in <module>
    for i in range(1,wd.max_row+1):
NameError: name 'wd' is not defined. Did you mean: 'wb'?
wb = openpyxl.load_workbook(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')
                         
for i in range(1,wb.max_row+1):
                         print(ws.cell(row=i,column=1).value)

                         
Traceback (most recent call last):
  File "<pyshell#15>", line 1, in <module>
    for i in range(1,wb.max_row+1):
AttributeError: 'Workbook' object has no attribute 'max_row'
wb=openpyxl.Workbook()
                         
wb = openpyxl.load_workbook(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')
                         
wb.max_row()
                         
Traceback (most recent call last):
  File "<pyshell#18>", line 1, in <module>
    wb.max_row()
AttributeError: 'Workbook' object has no attribute 'max_row'
for i in range(1,10):

                        print(ws.cell(row=i,column=1).value)

                         
None
None
None
None
None
None
None
None
None
wb.max_row()
                         
Traceback (most recent call last):
  File "<pyshell#22>", line 1, in <module>
    wb.max_row()
AttributeError: 'Workbook' object has no attribute 'max_row'
wswb.max_row()
                         
Traceback (most recent call last):
  File "<pyshell#23>", line 1, in <module>
    wswb.max_row()
NameError: name 'wswb' is not defined
ws=wb.active
                         
for i in range(1,10):

                        print(ws.cell(row=i,column=1).value)

                         
marketplace
US
US
US
US
US
US
US
US
for i in range(1,10):

                        print(ws.cell(row=1,column=i).value)

                         
marketplace
customer_id
review_id
product_id
product_parent
product_title
product_category
star_rating
helpful_votes
ws.max_column
                         
16
for i in range(1,ws.max_coumn+1):

                        print(ws.cell(row=1,column=i).value)

                         
Traceback (most recent call last):
  File "<pyshell#31>", line 1, in <module>
    for i in range(1,ws.max_coumn+1):
AttributeError: 'Worksheet' object has no attribute 'max_coumn'. Did you mean: 'max_column'?
for i in range(1,ws.max_column+1):

                        print(ws.cell(row=1,column=i).value)

                         
marketplace
customer_id
review_id
product_id
product_parent
product_title
product_category
star_rating
helpful_votes
total_votes
vine
verified_purchase
review_headline
review_body
review_date
filter_literal
headers=[]
                         
for i in range(1,ws.max_column+1):

                        headers.append(ws.cell(row=1,column=i).value)

                         
headers
                         
['marketplace', 'customer_id', 'review_id', 'product_id', 'product_parent', 'product_title', 'product_category', 'star_rating', 'helpful_votes', 'total_votes', 'vine', 'verified_purchase', 'review_headline', 'review_body', 'review_date', 'filter_literal']
ws = wb.create_sheet('3 start and below')
                         
ws = wb.create_sheet('4 start')
                         
ws = wb.create_sheet('5 start')
                         
ws.active('3 start and below')
                         
Traceback (most recent call last):
  File "<pyshell#41>", line 1, in <module>
    ws.active('3 start and below')
AttributeError: 'Worksheet' object has no attribute 'active'
ws = wb.create_sheet('3 start and below')
                         
ws3 = wb.create_sheet('3 start and below')
                         
ws4 = wb.create_sheet('4 start')
                         
ws5 = wb.create_sheet('5 start')
                         
ws3.active
                         
Traceback (most recent call last):
  File "<pyshell#46>", line 1, in <module>
    ws3.active
AttributeError: 'Worksheet' object has no attribute 'active'
ws3=wb.active
                         
for i in range(1,ws.max_column+1):
                         print(ws.cell(row=1,column=i).value)

                         
None
for i in range(1,h):
                         ws.cell(row=1,column=i).value

                         
Traceback (most recent call last):
  File "<pyshell#52>", line 1, in <module>
    for i in range(1,h):
NameError: name 'h' is not defined
header.length()
                         
Traceback (most recent call last):
  File "<pyshell#53>", line 1, in <module>
    header.length()
NameError: name 'header' is not defined. Did you mean: 'headers'?
header.size()
                         
Traceback (most recent call last):
  File "<pyshell#54>", line 1, in <module>
    header.size()
NameError: name 'header' is not defined. Did you mean: 'headers'?
len(header)
                         
Traceback (most recent call last):
  File "<pyshell#55>", line 1, in <module>
    len(header)
NameError: name 'header' is not defined. Did you mean: 'headers'?
len(headers)
                         
16
for i in range(0,len(headers)):
                         ws.cell(row=1,column=i+1).value=headers[i]

                         

for i in range(0,len(headers)):
                         print(ws.cell(row=1,column=i+1).value)

                         
marketplace
customer_id
review_id
product_id
product_parent
product_title
product_category
star_rating
helpful_votes
total_votes
vine
verified_purchase
review_headline
review_body
review_date
filter_literal
for i in range(0,len(headers)):
                         ws3.cell(row=1,column=i+1).value=headers[i]

                         
for i in range(0,len(headers)):
                         print(ws3.cell(row=1,column=i+1).value)

                         
marketplace
customer_id
review_id
product_id
product_parent
product_title
product_category
star_rating
helpful_votes
total_votes
vine
verified_purchase
review_headline
review_body
review_date
filter_literal
for i in headers:
                         ws4.append(i)

                         
Traceback (most recent call last):
  File "<pyshell#68>", line 2, in <module>
    ws4.append(i)
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/worksheet/worksheet.py", line 676, in append
    self._invalid_row(iterable)
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/worksheet/worksheet.py", line 812, in _invalid_row
    raise TypeError('Value must be a list, tuple, range or generator, or a dict. Supplied value is {0}'.format(
TypeError: Value must be a list, tuple, range or generator, or a dict. Supplied value is <class 'str'>
ws4.append(headers)
for i in range(0,len(headers)):
                         print(ws4.cell(row=1,column=i+1).value)

marketplace
customer_id
review_id
product_id
product_parent
product_title
product_category
star_rating
helpful_votes
total_votes
vine
verified_purchase
review_headline
review_body
review_date
filter_literal
ws5.append(headers)
wb.save(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')
data=[]
data=headers[0:2]
data
['marketplace', 'customer_id']
for i in range(1,wb.max_row+1):
if(ws.cell(row=1,column=1).value=='verified_purchase'):
SyntaxError: expected an indented block after 'for' statement on line 1
for i in range(1,wb.max_row+1):
    if(ws.cell(row=1,column=1).value=='verified_purchase'):
        for j in range(2,


                       u7u7
                       6
                       
SyntaxError: '(' was never closed
67
                       
67
for i in range(2,ws.max_row+1):
                       if(ws.cell(row=i,column=12).value=='Y') and int(ws.cell(row=i,column=8).value)<=3):
SyntaxError: unmatched ')'
for i in range(2,ws.max_row+1):
                       if(ws.cell(row=i,column=12).value=='Y') and int(ws.cell(row=i,column=8).value)<=3)):
        
SyntaxError: unmatched ')'
for i in range(2,ws.max_row+1):
                       if((ws.cell(row=i,column=12).value=='Y') and int(ws.cell(row=i,column=8).value)<=3)):
                         
SyntaxError: unmatched ')'
for i in range(2,ws.max_row+1):
                       if((ws.cell(row=i,column=12).value=='Y') and int(ws.cell(row=i,column=8).value)<=3))):
        
SyntaxError: unmatched ')'
for i in range(2,ws.max_row+1):
                       if(ws.cell(row=i,column=12).value=='Y') and int(ws.cell(row=i,column=8).value)<=3):
                           
SyntaxError: unmatched ')'
for i in range(2,ws.max_row+1):
                       if(ws.cell(row=i,column=12).value=='Y' and int(ws.cell(row=i,column=8).value)<=3):
                           for j in range(i,ws.max_column+1):
                               data=ws.cell(row=i,column=j).value
                               ws3.append(data)
                               ws3.save(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')

                               
data
['marketplace', 'customer_id']
for i in range(2,ws.max_row+1):
                       if(ws.cell(row=i,column=12).value=='Y' and int(ws.cell(row=i,column=8).value)<=3):
                           for j in range(i,ws.max_column+1):
                               data.append(ws.cell(row=i,column=j).value)
                               ws3.append(data)
                               ws3.save(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')

                               
data
['marketplace', 'customer_id']
data=[]
data
[]
for i in range(2,ws.max_row+1):
                       if(ws.cell(row=i,column=12).value=='Y' and int(ws.cell(row=i,column=8).value)<=3):
                           for j in range(i,ws.max_column+1):
                               data.append(ws.cell(row=i,column=j).value)
                               ws3.append(data)
                               ws3.save(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')

                               

data
[]
for i in range(2,ws.max_row+1):
                       if(ws.cell(row=i,column=12).value=='Y' and int(ws.cell(row=i,column=8).value)==4):
                           for j in range(i,ws.max_column+1):
                               data.append(ws.cell(row=i,column=j).value)
                               ws4.append(data)
                               ws4.save(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')

                               
data
[]
for i in range(2,ws.max_row+1):
                       if(ws.cell(row=i,column=12).value=='Y' and int(ws.cell(row=i,column=8).value)==4):
                           for j in range(i,ws.max_column+1):
                               data.append(ws.cell(row=i,column=j).value)
                               ws4.append(data)
                               ws4.save(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')

                               
for i in range(2,ws.max_row+1):
                       if(ws.cell(row=i,column=12).value=='Y' and int(ws.cell(row=i,column=8).value)==4):
                           for j in range(i,ws.max_column+1):
                               data.append(ws.cell(row=i,column=j).value)
                               ws4.append(data)
                               ws4.save(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')

                               
data
[]
ws=wb.active
for i in range(2,ws.max_row+1):
    print('for loop 1')
                       if(ws.cell(row=i,column=12).value=='Y' and int(ws.cell(row=i,column=8).value)==4):
                           print('into if statement')
                           for j in range(i,ws.max_column+1):
                               print('into for loop')
                               data.append(ws.cell(row=i,column=j).value)
                               ws4.append(data)
                               ws4.save(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')
                               
SyntaxError: unexpected indent
for i in range(2,ws.max_row+1):
                       print('for loop 1')
                       if(ws.cell(row=i,column=12).value=='Y' and int(ws.cell(row=i,column=8).value)==4):
                           print('into if statement')
                           for j in range(i,ws.max_column+1):
                               print('into for loop')
                               data.append(ws.cell(row=i,column=j).value)
                               ws4.append(data)
                               ws4.save(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')

                               
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
into if statement
into for loop
Traceback (most recent call last):
  File "<pyshell#120>", line 9, in <module>
    ws4.save(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')
AttributeError: 'Worksheet' object has no attribute 'save'
for i in range(2,ws.max_row+1):
                       print('for loop 1')
                       if(ws.cell(row=i,column=12).value=='Y' and int(ws.cell(row=i,column=8).value)==4):
                           print('into if statement')
                           for j in range(i,ws.max_column+1):
                               print('into for loop')
                               data.append(ws.cell(row=i,column=j).value)
                               ws4.append(data)
                               wb.save(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')

                               
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
into if statement
into for loop
into for loop
into for loop
into for loop
into for loop
into for loop
into for loop
into for loop
into for loop
into for loop
into for loop
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
into if statement
into for loop
into for loop
into for loop
into for loop
into for loop
into for loop
for loop 1
into if statement
into for loop
into for loop
into for loop
into for loop
into for loop
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
into if statement
for loop 1
for loop 1
for loop 1
into if statement
for loop 1
into if statement
for loop 1
into if statement
for loop 1
into if statement
for loop 1
for loop 1
for loop 1
for loop 1
into if statement
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
into if statement
for loop 1
for loop 1
for loop 1
for loop 1
into if statement
for loop 1
for loop 1
for loop 1
for loop 1
into if statement
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
into if statement
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
into if statement
for loop 1
for loop 1
into if statement
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
into if statement
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
into if statement
for loop 1
for loop 1
for loop 1
for loop 1
into if statement
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
into if statement
for loop 1
for loop 1
for loop 1
for loop 1
for loop 1
into if statement
for i in range(2,ws.max_row+1):
                       if(ws.cell(row=i,column=12).value=='Y' and int(ws.cell(row=i,column=8).value)==5):
                           for j in range(i,ws.max_column+1):
                               data.append(ws.cell(row=i,column=j).value)
                               ws5.append(data)
                               wb.save(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')

                               



ws4 = wb.create_sheet('4 start')
wb.save(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')
ws4.append(headers)
headers
['marketplace', 'customer_id', 'review_id', 'product_id', 'product_parent', 'product_title', 'product_category', 'star_rating', 'helpful_votes', 'total_votes', 'vine', 'verified_purchase', 'review_headline', 'review_body', 'review_date', 'filter_literal']
for i in range(2,ws.max_row+1):
                       if(ws.cell(row=i,column=12).value=='Y' and int(ws.cell(row=i,column=8).value)==4):
                           for j in range(i,ws.max_column+1):
                               data.append(ws.cell(row=i,column=j).value)
                               ws4.append(data)
                               wb.save(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')

                               

wb = openpyxl.load_workbook(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')
Traceback (most recent call last):
  File "<pyshell#132>", line 1, in <module>
    wb = openpyxl.load_workbook(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/reader/excel.py", line 317, in load_workbook
    reader.read()
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/reader/excel.py", line 276, in read
    self.read_manifest()
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/reader/excel.py", line 134, in read_manifest
    src = self.archive.read(ARC_CONTENT_TYPES)
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/zipfile.py", line 1473, in read
    with self.open(name, "r", pwd) as fp:
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/zipfile.py", line 1512, in open
    zinfo = self.getinfo(name)
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/zipfile.py", line 1439, in getinfo
    raise KeyError(
KeyError: "There is no item named '[Content_Types].xml' in the archive"
wb = openpyxl.load_workbook(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')
Traceback (most recent call last):
  File "<pyshell#133>", line 1, in <module>
    wb = openpyxl.load_workbook(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/reader/excel.py", line 317, in load_workbook
    reader.read()
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/reader/excel.py", line 276, in read
    self.read_manifest()
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/reader/excel.py", line 134, in read_manifest
    src = self.archive.read(ARC_CONTENT_TYPES)
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/zipfile.py", line 1473, in read
    with self.open(name, "r", pwd) as fp:
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/zipfile.py", line 1512, in open
    zinfo = self.getinfo(name)
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/zipfile.py", line 1439, in getinfo
    raise KeyError(
KeyError: "There is no item named '[Content_Types].xml' in the archive"
wb = openpyxl.load_workbook(filename='/Users/sriram/Downloads/sample.xlsx')
Traceback (most recent call last):
  File "<pyshell#134>", line 1, in <module>
    wb = openpyxl.load_workbook(filename='/Users/sriram/Downloads/sample.xlsx')
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/reader/excel.py", line 317, in load_workbook
    reader.read()
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/reader/excel.py", line 276, in read
    self.read_manifest()
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages/openpyxl/reader/excel.py", line 134, in read_manifest
    src = self.archive.read(ARC_CONTENT_TYPES)
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/zipfile.py", line 1473, in read
    with self.open(name, "r", pwd) as fp:
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/zipfile.py", line 1512, in open
    zinfo = self.getinfo(name)
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/zipfile.py", line 1439, in getinfo
    raise KeyError(
KeyError: "There is no item named '[Content_Types].xml' in the archive"
wb = openpyxl.load_workbook(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')
headers
['marketplace', 'customer_id', 'review_id', 'product_id', 'product_parent', 'product_title', 'product_category', 'star_rating', 'helpful_votes', 'total_votes', 'vine', 'verified_purchase', 'review_headline', 'review_body', 'review_date', 'filter_literal']
ws4=wb.create_sheet('4 start')
data
["Orient ER27009B Men's Symphony Automatic Stainless Steel Black Dial Mechanical Watch", "Orient ER27009B Men's Symphony Automatic Stainless Steel Black Dial Mechanical Watch", 'Watches', 4, 0, 0, 'N', 'Y', 'Beautiful face, but cheap sounding links', "Beautiful watch face.  The band looks nice all around.  The links do make that squeaky cheapo noise when you swing it back and forth on your wrist which can be embarrassing in front of watch enthusiasts.  However, to the naked eye from afar, you can't tell the links are cheap or folded because it is well polished and brushed and the folds are pretty tight for the most part.<br /><br />I love the new member of my collection and it looks great.  I've had it for about a week and so far it has kept good time despite day 1 which is typical of a new mechanical watch", '2015-08-31', 'Add', 'N', 'Y', 'Great quality and build', 'Great quality and build.<br />The motors are really silent.<br />After fiddling with the settings my watches are always charged and ready to use.', '2015-08-31', 'Add', 'Y', 'Satisfied', "The watch was pretty much as it was described and how it looks. I really like the simplicity of it and it looks pretty amazing. It is slightly too big, so I have yet to adjust it, but I'm glad it's too big rather than too small. The weight is good which is what I was hoping for. Very satisfied with this purchase. I was slightly on the fence getting this one because of the previous fossil watch I had, which by some time, did have some gems fall off from ring around the numbers, but so far so good with this one! Excuse my bad lighting for my pic.", '2015-08-31', 'Add', 3653882, 'R3O9SGZBVQBV76', 'B00FALQ1ZC', 937001370, 'Invicta Women\'s 15150 "Angel" 18k Yellow Gold Ion-Plated Stainless Steel and Brown Leather Watch', 'Watches', 5, 0, 0, 'N', 'Y', 'Five Stars', 'Absolutely love this watch! Get compliments almost every time I wear it. Dainty.', '2015-08-31', 'Add', 'RKH8BNC3L5DLF', 'B00D3RGO20', 484010722, "Kenneth Cole New York Women's KC4944 Automatic Silver Automatic Mesh Bracelet Analog Watch", 'Watches', 5, 0, 0, 'N', 'Y', 'I love thiswatch it keeps time wonderfully', 'I love this watch it keeps time wonderfully.', '2015-08-31', None, 958035625, "Citizen Men's BM8180-03E Eco-Drive Stainless Steel Watch with Green Canvas Band", 'Watches', 5, 0, 0, 'N', 'Y', 'Five Stars', 'It works well on me. However, I found cheaper prices in other places after making the purchase', '2015-08-31', 'Add', 'Watches', 5, 0, 0, 'N', 'Y', 'No complaints', 'i love this watch for my purpose, about the people complaining should of done their research better before buying. dumb people.', '2015-08-31', 'Add', 5, 1, 1, 'N', 'Y', 'Five Stars', 'for my wife and she loved it, looks great and a great price!', '2015-08-31', 'Add', 2, 'N', 'Y', 'Perfect watch!', "Watch is perfect. Rugged with the metal &#34;Bull Bars&#34;. The red accents are a great touch and I get compliments when wearing it. If you are worried about being able to read this in sunlight or in the dark don't! The LED ilumination works great! I might even get this in a different color for my next G-Shock purchase!", '2015-08-31', 'Add', '2015-08-31', None, None, "Orient ER27009B Men's Symphony Automatic Stainless Steel Black Dial Mechanical Watch", 'Watches', 4, 0, 0, 'N', 'Y', 'Beautiful face, but cheap sounding links', "Beautiful watch face.  The band looks nice all around.  The links do make that squeaky cheapo noise when you swing it back and forth on your wrist which can be embarrassing in front of watch enthusiasts.  However, to the naked eye from afar, you can't tell the links are cheap or folded because it is well polished and brushed and the folds are pretty tight for the most part.<br /><br />I love the new member of my collection and it looks great.  I've had it for about a week and so far it has kept good time despite day 1 which is typical of a new mechanical watch", '2015-08-31', 'Add', 'N', 'Y', 'Great quality and build', 'Great quality and build.<br />The motors are really silent.<br />After fiddling with the settings my watches are always charged and ready to use.', '2015-08-31', 'Add', 'Y', 'Satisfied', "The watch was pretty much as it was described and how it looks. I really like the simplicity of it and it looks pretty amazing. It is slightly too big, so I have yet to adjust it, but I'm glad it's too big rather than too small. The weight is good which is what I was hoping for. Very satisfied with this purchase. I was slightly on the fence getting this one because of the previous fossil watch I had, which by some time, did have some gems fall off from ring around the numbers, but so far so good with this one! Excuse my bad lighting for my pic.", '2015-08-31', 'Add']
data=[]
for i in range(2,ws.max_row+1):
                       if(ws.cell(row=i,column=12).value=='Y' and int(ws.cell(row=i,column=8).value)==4):
                           for j in range(i,ws.max_column+1):
                               data=[]
                               data.append(ws.cell(row=i,column=j).value)
                               ws4.append(data)
                               wb.save(filename='/Users/sriram/Downloads/openpyxl-sample.xlsx')

ws4 = wb.create_sheet('4 start')
ws4.append(headers)
res=Request('https://randomuser.me/api/?results=2000')
Traceback (most recent call last):
  File "<pyshell#144>", line 1, in <module>
    res=Request('https://randomuser.me/api/?results=2000')
NameError: name 'Request' is not defined
res=request('https://randomuser.me/api/?results=2000')
Traceback (most recent call last):
  File "<pyshell#145>", line 1, in <module>
    res=request('https://randomuser.me/api/?results=2000')
NameError: name 'request' is not defined
import requests
res=request('https://randomuser.me/api/?results=2000')
Traceback (most recent call last):
  File "<pyshell#147>", line 1, in <module>
    res=request('https://randomuser.me/api/?results=2000')
NameError: name 'request' is not defined. Did you mean: 'requests'?
res=requests('https://randomuser.me/api/?results=2000')
Traceback (most recent call last):
  File "<pyshell#148>", line 1, in <module>
    res=requests('https://randomuser.me/api/?results=2000')
TypeError: 'module' object is not callable
res=requests.get('https://randomuser.me/api/?results=2000')
res.status_code
200
type(res)
<class 'requests.models.Response'>
import json
datapy=json.loads(res)
Traceback (most recent call last):
  File "<pyshell#153>", line 1, in <module>
    datapy=json.loads(res)
  File "/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/json/__init__.py", line 339, in loads
    raise TypeError(f'the JSON object must be str, bytes or bytearray, '
TypeError: the JSON object must be str, bytes or bytearray, not Response
data=res.text
                    
type(data)
                    
<class 'str'>
len(data)
                    
2159861
datapy=json.loads(data)
                    
len(datapy)
                    
2
datapy
                    








ssd
type(datapy)
                    
<class 'dict'>
datapy.keys
                    
<built-in method keys of dict object at 0x11712c740>
datapy.keys()
                    
dict_keys(['results', 'info'])
len(datapy['results'])
                    
2000
headers=['First','Last','Gender','City','email','age','phone','cell']
                    
wb=openpyxl.workbook()
                    
Traceback (most recent call last):
  File "<pyshell#165>", line 1, in <module>
    wb=openpyxl.workbook()
TypeError: 'module' object is not callable
import openpyxl
wb=openpyxl.Workbook()
ws=wb.active
ws.append(headers)
resList=datapy['results']
len(resList[0])
12
resList[0]['first']
Traceback (most recent call last):
  File "<pyshell#172>", line 1, in <module>
    resList[0]['first']
KeyError: 'first'
resList[0]
{'gender': 'male', 'name': {'title': 'Mr', 'first': 'Bing', 'last': 'Van Hooren'}, 'location': {'street': {'number': 2807, 'name': 'Kronenburggalerij'}, 'city': 'Mander', 'state': 'Groningen', 'country': 'Netherlands', 'postcode': '3799 PX', 'coordinates': {'latitude': '38.5579', 'longitude': '-1.4296'}, 'timezone': {'offset': '-11:00', 'description': 'Midway Island, Samoa'}}, 'email': 'bing.vanhooren@example.com', 'login': {'uuid': '1d04f86e-6dd5-4535-b9f6-7913f7943780', 'username': 'yellowgorilla546', 'password': 'mature', 'salt': 'z13wxnxS', 'md5': '4dc0b87b5d6c5d43e38b00f87750c80c', 'sha1': 'c55c3b69c75252db9ed9e73f454cb548bf9d83bd', 'sha256': '94ab0e83017e225677e6e843b1bbb6ba49bb0b0bc8609a52f2b2e9d49a296a65'}, 'dob': {'date': '1979-05-21T13:19:41.371Z', 'age': 43}, 'registered': {'date': '2004-05-13T11:00:44.527Z', 'age': 18}, 'phone': '(094) 5224506', 'cell': '(06) 36469831', 'id': {'name': 'BSN', 'value': '83272727'}, 'picture': {'large': 'https://randomuser.me/api/portraits/men/20.jpg', 'medium': 'https://randomuser.me/api/portraits/med/men/20.jpg', 'thumbnail': 'https://randomuser.me/api/portraits/thumb/men/20.jpg'}, 'nat': 'NL'}
resList[0]['name']['first']
'Bing'
resList[0]['name']['last']
'Van Hooren'
resList['name']['first']
Traceback (most recent call last):
  File "<pyshell#176>", line 1, in <module>
    resList['name']['first']
TypeError: list indices must be integers or slices, not str
resList[0]['gender']
'male'
for i in range(len(results)):
    listData=[]
    listData.append(resList[0]['name']['first'])
    listData.append(resList[0]['name']['last'])

    
Traceback (most recent call last):
  File "<pyshell#182>", line 1, in <module>
    for i in range(len(results)):
NameError: name 'results' is not defined. Did you mean: 'requests'?
for i in range(len(datapy['results'])):
    listData=[]
    listData.append(resList[0]['name']['first'])
    listData.append(resList[0]['name']['last'])
    listData.append(resList[0]['gender'])
    listData.append(resList[0]['location']['city'])
    listData.append(resList[0]['email'])
    listData.append(resList[0]['dob']['age'])
    ws.append(listData)
    wd.save(filename='randomUSerData.xlsx')

    
Traceback (most recent call last):
  File "<pyshell#190>", line 10, in <module>
    wd.save(filename='randomUSerData.xlsx')
NameError: name 'wd' is not defined. Did you mean: 'wb'?
for i in range(len(datapy['results'])):
    listData=[]
    listData.append(resList[0]['name']['first'])
    listData.append(resList[0]['name']['last'])
    listData.append(resList[0]['gender'])
    listData.append(resList[0]['location']['city'])
    listData.append(resList[0]['email'])
    listData.append(resList[0]['dob']['age'])
    ws.append(listData)
    wb.save(filename='randomUSerData.xlsx')

                    







listData
break;


import os
os.getcwd()
'/Users/sriram/Documents'
wb=openpyxl.Workbook()
ws=wb.active
len(resList[0])
12
ws.append(headers)
headers
['First', 'Last', 'Gender', 'City', 'email', 'age', 'phone', 'cell']
wb.save(filename='randomUSerData.xlsx')
\
SyntaxError: multiple statements found while compiling a single statement
wb.save(filename='randomData.xlsx')
resList[0]['name']['first']
'Bing'
resList[0]['name']['first']
'Bing'
resList[0]['name']['first']
'Bing'
resList['name']['first']
Traceback (most recent call last):
  File "<pyshell#205>", line 1, in <module>
    resList['name']['first']
TypeError: list indices must be integers or slices, not str
for i in range(len(datapy['results'])):
    listData=[]
    listData.append(resList[i]['name']['first'])
    listData.append(resList[i]['name']['last'])
    listData.append(resList[i]['gender'])
    listData.append(resList[i]['location']['city'])
    listData.append(resList[i]['email'])
    listData.append(resList[i]['dob']['age'])
    listData.append(resList[i]['phone'])
    listData.append(resList[i]['cell'])
    ws.append(listData)
    wb.save(filename='randomData.xlsx')

    
